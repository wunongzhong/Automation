# 作者：daniel
# @Time: 2020/8/12 19:12
# @File: conftest.PY
import pytest
import requests
import json
import os
from openpyxl import workbook,load_workbook
from tools.log import logger
from requestcommon.configcommon import test_case_path

@pytest.fixture(scope='session')
def get_excel():
    logger.info("遍历用例目录读取用例excel开始。。。。。。。。。。。。。")
    file_name_list = os.walk(test_case_path, topdown=True)
    dict = {}

    for root, dirs, file_name in file_name_list:
        try:
            for name in file_name:
                print(os.path.join(root, name))
                logger.info("读取用例excel中sheet开始。。。。。。。。。。。。。")
                workbook = load_workbook(os.path.join(root, name))
                sheet = workbook['Sheet1']

                lists = []
                rows_sheet = sheet.iter_rows()
                for item in rows_sheet:
                    if item[0].value == "api":
                        continue
                    list = []
                    for col in item:
                        logger.info("遍历每一列加入到一行测试数据。。。。。。。。。。。。。")
                        list.append(col.value)
                    lists.append(list)
                dict[name] = lists
        except Exception as e:
            logger.error("历用例目录读取用例excel执行出错，请查看问题！原因: s%", e)
    #print(dict)
    return dict
