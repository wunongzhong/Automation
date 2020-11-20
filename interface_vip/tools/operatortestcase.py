# -*- coding: utf-8 -*-
# @Time     : 2020/07/02
# @Author   : daniel
# @File     : operatortestcase.py
# @Desc     : 循环读取用例目录
# @Version  : V1.0.0
import os

import openpyxl
from openpyxl import load_workbook

from requestcommon.configcommon import test_case_path


class readCaseFromDir:

    # 加载测试用例
    def get_test_case(file_name):
        testcase_data = []
        test_case_file = os.path.join(test_case_path, file_name)
        wb = load_workbook(test_case_file)
        for test_case_sheet in wb.sheetnames:
            excel = ParseExcel(test_case_file, test_case_sheet)
            testcase_data = testcase_data + (excel.get_testcase_from_sheet(test_case_file, test_case_sheet))
        return testcase_data



