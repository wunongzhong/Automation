#coding=utf-8
import openpyxl
from openpyxl.styles import PatternFill,Font
from KeyWords_demo.web_keywords import WebUIKeys

def read_excel():
    excel = openpyxl.load_workbook('yongli.xlsx')
    for i in excel.sheetnames:
        sheet = excel[i]
        #读取所有内容
        for value in sheet.values:
            if type(value[0]) is int:
                parmDict = dict.fromkeys(('name','value','txt','expect'))
                parmDict['name'] = value[2]
                parmDict['value'] = value[3]
                parmDict['txt'] = value[4]
                parmDict['expect'] = value[6]
                if value[1] == 'open_brower':
                    wk = WebUIKeys(value[4])
                elif 'assert' in value[1]:
                    a = getattr(wk,value[1])(**parmDict)
                    row = value[0] + 1
                    if a:
                        sheet.cell(row=row, column=8).value = 'Pass'
                        sheet.cell(row=row, column=8).fill = PatternFill('solid', fgColor='00FF00')
                        sheet.cell(row=row, column=8).font = Font(bold=True)
                    else:
                        sheet.cell(row=row, column=8).value = 'False'
                        sheet.cell(row=row, column=8).fill = PatternFill('solid', fgColor='FF0000')
                        sheet.cell(row=row, column=8).font = Font(bold=True)
                elif value[1] == 'quit' or value[1] == 'close':
                    getattr(wk, value[1])()
                else:
                    getattr(wk,value[1])(**parmDict)
    excel.save('yongli.xlsx')
    excel.close()

# def read_test():
#     excel = openpyxl.load_workbook('yongli.xlsx')
#     for i in excel.sheetnames:
#         sheet = excel[i]
#         # 读取所有内容
#         for value in sheet.values:
#             print(value)
#     excel.close()

if __name__ == '__main__':
    read_excel()